from collections import Counter
import sys
from openpyxl import load_workbook


def find_duplicate_emails(file_path: str, sheet_name: str = None) -> None:
    # Load workbook
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    # Read header row
    headers = [cell.value for cell in sheet[1]]

    if "Email" not in headers:
        print('Error: "Email" column not found.')
        return

    email_col_index = headers.index("Email") + 1  # openpyxl is 1-based

    emails = []
    row_map = {}

    # Read email values from rows
    for row_num in range(2, sheet.max_row + 1):
        cell_value = sheet.cell(row=row_num, column=email_col_index).value

        if cell_value is None:
            continue

        email = str(cell_value).strip().lower()
        if not email:
            continue

        emails.append(email)
        row_map.setdefault(email, []).append(row_num)

    # Count duplicates
    counts = Counter(emails)
    duplicates = {email: count for email, count in counts.items() if count > 1}

    if not duplicates:
        print("No duplicate emails found.")
        return

    print("Duplicate email report:")
    for email, count in duplicates.items():
        print(f'"{email}" is being duplicate ({count} times) in rows {row_map[email]}')


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python find_duplicate_emails.py <xlsx_file_path> [sheet_name]")
        sys.exit(1)

    file_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None

    find_duplicate_emails(file_path, sheet_name)