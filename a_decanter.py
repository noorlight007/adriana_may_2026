import os
from dotenv import load_dotenv
load_dotenv()

import re
import json
from typing import Any, Dict

from openai import OpenAI
from jsonschema import Draft202012Validator
from openpyxl import load_workbook


MODEL = "gpt-4o-search-preview-2025-03-11"
EXCEL_FILE = "awards_decanter_France.xlsx"
COUNTRY_NAME = "France"


def return_schema() -> Dict[str, Any]:
    schema: Dict[str, Any] = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "website": {
                "type": "string",
                "description": 'Official company website URL. Return "null" if not found.'
            },
            "email": {
                "type": "string",
                "description": 'Employee/company work email from the found website. Return "null" if not found.'
            }
        },
        "required": ["website", "email"]
    }
    return schema


def get_all_instructions():
    system_instructions = (
        "You are an information extraction engine.\n"
        "Your task is to find the official website of a company and then extract one email address from that website.\n"
        "- The company is located in France.\n"
        "- First, identify the most likely official website for the company.\n"
        "- Only use a website if you are confident it is the company's real official website.\n"
        "- If you cannot find the official website, return \"null\" for both website and email.\n"
        "- If website is not found, do NOT guess or infer an email.\n"
        "- If website is found, then try to extract one email address explicitly available on that website.\n"
        "- Prefer an employee's/company work email if clearly available.\n"
        "- If no employee work email is found, use the company's general contact email if explicitly available.\n"
        "- Do NOT invent, guess, or infer any website or email address.\n"
        "- Only extract information explicitly supported by search/website content.\n"
        "- Return ONLY raw JSON matching the required schema.\n"
        "- Do not return markdown fences, explanations, or extra text.\n"
        "- Output format must be exactly like:\n"
        "{\n"
        "  \"website\": \"https://example.com\",\n"
        "  \"email\": \"info@example.com\"\n"
        "}\n"
        "- If website is found but email is not found, output:\n"
        "{\n"
        "  \"website\": \"https://example.com\",\n"
        "  \"email\": \"null\"\n"
        "}\n"
        "- If website is not found, output:\n"
        "{\n"
        "  \"website\": \"null\",\n"
        "  \"email\": \"null\"\n"
        "}\n"
    )

    user_instructions_template = (
        "Find the official website and one relevant email address for the following company.\n"
        "Country: {country_name}\n"
        "Company Name: {company_name}\n\n"
        "Rules:\n"
        "- Do not invent anything.\n"
        "- If you do not find the official website, return \"null\" for both website and email.\n"
        "- If you do find the official website, then try to find an email from that website.\n"
        "- Prefer an employee's company email.\n"
        "- If no employee email is available, return a general company email if explicitly shown.\n"
        "- If no email is found on the website, return \"null\" for email.\n"
        "- Return ONLY raw JSON matching the required schema.\n"
    )

    return system_instructions, user_instructions_template


def strip_code_fences(s: str) -> str:
    s = s.strip()
    if s.startswith("```"):
        lines = s.splitlines()
        if lines and lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        s = "\n".join(lines).strip()
    return s


def find_json_block(s: str) -> str:
    s = s.strip()
    obj_match = re.search(r"\{.*\}\s*$", s, flags=re.DOTALL)
    if obj_match:
        return obj_match.group(0)
    arr_match = re.search(r"\[.*\]\s*$", s, flags=re.DOTALL)
    if arr_match:
        return arr_match.group(0)
    return s


def normalize_nulls(data: Dict[str, Any]) -> Dict[str, Any]:
    def to_str_null(v: Any) -> str:
        if v is None:
            return "null"
        if isinstance(v, str):
            cleaned = v.strip()
            return cleaned if cleaned else "null"
        return str(v)

    data["website"] = to_str_null(data.get("website"))
    data["email"] = to_str_null(data.get("email"))
    return data


def get_header_indexes(ws) -> Dict[str, int]:
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is not None:
            headers[str(value).strip()] = col
    return headers


def extract_website_and_email(client: OpenAI, company_name: str, country_name: str) -> Dict[str, str]:
    schema = return_schema()
    system_instructions, user_instructions_template = get_all_instructions()

    messages = [
        {"role": "system", "content": system_instructions},
        {
            "role": "user",
            "content": user_instructions_template.format(
                company_name=company_name,
                country_name=country_name,
            ),
        },
    ]

    resp = client.chat.completions.create(
        model=MODEL,
        messages=messages,
    )

    raw = resp.choices[0].message.content or ""
    content = strip_code_fences(raw)

    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        candidate = find_json_block(content)
        data = json.loads(candidate)

    Draft202012Validator(schema).validate(data)
    data = normalize_nulls(data)

    return {
        "website": data["website"],
        "email": data["email"],
    }


def main():
    api_key = os.getenv("OPENAI_API")
    if not api_key:
        raise ValueError("OPENAI_API environment variable is not set.")

    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"{EXCEL_FILE} not found.")

    client = OpenAI(api_key=api_key)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    headers = get_header_indexes(ws)

    if "Company Name" not in headers:
        raise ValueError(f'Column "Company Name" not found in {EXCEL_FILE}')
    if "Website" not in headers:
        raise ValueError(f'Column "Website" not found in {EXCEL_FILE}')
    if "Email" not in headers:
        raise ValueError(f'Column "Email" not found in {EXCEL_FILE}')

    company_name_col = headers["Company Name"]
    website_col = headers["Website"]
    email_col = headers["Email"]

    for row in range(2, ws.max_row + 1):
        company_name_value = ws.cell(row=row, column=company_name_col).value
        current_website = ws.cell(row=row, column=website_col).value
        current_email = ws.cell(row=row, column=email_col).value

        company_name = str(company_name_value).strip() if company_name_value else ""
        existing_website = str(current_website).strip() if current_website else ""
        existing_email = str(current_email).strip() if current_email else ""

        if not company_name:
            print(f"Row {row}: No company name found. Skipping.")
            continue

        if existing_website or existing_email:
            print(
                f"Row {row}: Already has data "
                f"(Website: {existing_website or 'empty'}, Email: {existing_email or 'empty'}). Skipping."
            )
            continue

        try:
            result = extract_website_and_email(
                client=client,
                company_name=company_name,
                country_name=COUNTRY_NAME,
            )

            website = result["website"]
            email = result["email"]

            if website != "null":
                ws.cell(row=row, column=website_col).value = website
                print(f"Row {row}: Website written -> {website}")
            else:
                print(f"Row {row}: Website not found for {company_name}")

            if website != "null" and email != "null":
                ws.cell(row=row, column=email_col).value = email
                print(f"Row {row}: Email written -> {email}")
            elif website != "null":
                print(f"Row {row}: No email found on website for {company_name}")

            # Save after each processed row before moving to the next one
            wb.save(EXCEL_FILE)
            print(f"Row {row}: Saved to workbook.")
            print("--------------------------------------------------")

        except Exception as e:
            # Save workbook state even if this row fails, then continue
            wb.save(EXCEL_FILE)
            print(f"Row {row}: Error processing {company_name}: {e}")
            print("--------------------------------------------------")
            continue

    wb.save(EXCEL_FILE)
    wb.close()
    print("Done.")


if __name__ == "__main__":
    main()