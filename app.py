import json
import os
import re
import time
from typing import Any, Dict, List, Optional

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook


JSON_FILE = "aust_wi_list_id.json"
XLSX_FILE = "austwinelist.xlsx"

BASE_URL = "https://www.austrianwine.com/wines-wineries/winery"
REQUEST_TIMEOUT = 30
SLEEP_BETWEEN_REQUESTS = 0.5

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/122.0.0.0 Safari/537.36"
    )
}

# Expected Excel columns
EXPECTED_COLUMNS = [
    "Company name",
    "Email",
    "First name",
    "Last name",
    "Job title",
    "country/address",
    "Phone",
    "Website",
    "Industry type",
]


def safe_strip(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def normalize_website(url: Optional[str]) -> Optional[str]:
    if not url:
        return None
    url = url.strip()
    if not url:
        return None
    if not re.match(r"^https?://", url, flags=re.IGNORECASE):
        return "http://" + url
    return url


def load_input_json(path: str) -> List[Dict[str, Any]]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        raise ValueError(f"{path} must contain a JSON array")

    return data


def get_json_detail(winery_id: str) -> Optional[Dict[str, Any]]:
    """
    Try the json=true endpoint first because it is more stable and easier to parse.
    """
    params = {
        "json": "true",
        "tx_wineapi_wineriesdetail[winery]": winery_id,
    }

    try:
        response = requests.get(
            BASE_URL,
            params=params,
            headers=HEADERS,
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()

        content_type = response.headers.get("Content-Type", "")
        text = response.text.strip()

        if "application/json" in content_type.lower() or text.startswith("{"):
            return response.json()

    except Exception as e:
        print(f"[WARN] JSON request failed for {winery_id}: {e}")

    return None


def get_html_detail(winery_id: str) -> Optional[str]:
    """
    Fallback to the HTML page if JSON is unavailable.
    """
    params = {
        "tx_wineapi_wineriesdetail[winery]": winery_id,
    }

    try:
        response = requests.get(
            BASE_URL,
            params=params,
            headers=HEADERS,
            timeout=REQUEST_TIMEOUT,
        )
        response.raise_for_status()
        return response.text
    except Exception as e:
        print(f"[WARN] HTML request failed for {winery_id}: {e}")
        return None


def parse_from_json(data: Dict[str, Any]) -> Dict[str, Optional[str]]:
    """
    Parse fields from the JSON endpoint.
    """
    company_name = safe_strip(data.get("title"))
    phone = safe_strip(data.get("phone"))
    email = safe_strip(data.get("email"))

    website = (
        safe_strip(data.get("website"))
        or safe_strip(data.get("url"))
        or safe_strip(data.get("web"))
    )
    website = normalize_website(website)

    address = None
    location_data = data.get("address")

    if isinstance(location_data, str):
        address = safe_strip(location_data)
    elif isinstance(location_data, dict):
        # Try some common address structures
        parts = [
            safe_strip(location_data.get("street")),
            safe_strip(location_data.get("zip")),
            safe_strip(location_data.get("city")),
            safe_strip(location_data.get("country")),
        ]
        parts = [p for p in parts if p]
        address = ", ".join(parts) if parts else None

    return {
        "company_name": company_name,
        "address": address,
        "phone": phone,
        "email": email,
        "website": website,
    }


def parse_from_html(html: str) -> Dict[str, Optional[str]]:
    """
    Parse fields from HTML.

    Rules:
    - company name from <h1><mark>...</mark></h1>
    - phone from href starting with tel:
    - email from href starting with mailto:
    - website from normal external link in infobox
    - address from Google Maps link or first relevant row in infobox
    """
    soup = BeautifulSoup(html, "html.parser")

    company_name = None
    address = None
    phone = None
    email = None
    website = None

    # Company name: <h1><mark>...</mark></h1>
    h1_mark = soup.select_one("h1 mark")
    if h1_mark:
        company_name = safe_strip(h1_mark.get_text(" ", strip=True))
    else:
        h1 = soup.find("h1")
        if h1:
            company_name = safe_strip(h1.get_text(" ", strip=True))

    infobox = soup.select_one("div.infobox")

    if infobox:
        links = infobox.find_all("a", href=True)

        for a in links:
            href = a["href"].strip()
            text = safe_strip(a.get_text(" ", strip=True))

            href_lower = href.lower()

            if href_lower.startswith("tel:") and phone is None:
                phone = text or href[4:].strip()

            elif href_lower.startswith("mailto:") and email is None:
                email = text or href[7:].strip()

            elif "maps.google.com" in href_lower and address is None:
                address = text

            elif website is None:
                # External website link, but not maps/mail/tel
                if href_lower.startswith(("http://", "https://")):
                    if "maps.google.com" not in href_lower:
                        website = normalize_website(href)

        # Extra fallback for address if maps link not found
        if address is None:
            rows = infobox.select("div.row")
            for row in rows:
                row_text = safe_strip(row.get_text(" ", strip=True))
                if row_text and not any(
                    token in row_text.lower()
                    for token in ["@", "http", "www", "+43", "tel", "phone"]
                ):
                    address = row_text
                    break

    return {
        "company_name": company_name,
        "address": address,
        "phone": phone,
        "email": email,
        "website": website,
    }


def fetch_winery_data(winery_id: str) -> Dict[str, Optional[str]]:
    """
    Fetch winery detail, prefer JSON endpoint, fallback to HTML parsing.
    """
    result = {
        "company_name": None,
        "address": None,
        "phone": None,
        "email": None,
        "website": None,
    }

    json_data = get_json_detail(winery_id)
    if json_data:
        parsed = parse_from_json(json_data)
        result.update(parsed)

        # If JSON misses some fields, try HTML as fallback
        if not all([result["company_name"], result["phone"], result["email"], result["website"]]):
            html = get_html_detail(winery_id)
            if html:
                html_parsed = parse_from_html(html)
                for key, value in html_parsed.items():
                    if not result.get(key) and value:
                        result[key] = value
        return result

    html = get_html_detail(winery_id)
    if html:
        return parse_from_html(html)

    return result


def open_or_create_workbook(path: str):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(EXPECTED_COLUMNS)
        wb.save(path)

    # Ensure header exists and map columns
    headers = [cell.value for cell in ws[1]]

    if headers != EXPECTED_COLUMNS:
        # Build a header map from existing headers if possible
        # If sheet is empty or malformed, fix only the first row
        if ws.max_row == 1 and all(v is None for v in headers):
            for idx, name in enumerate(EXPECTED_COLUMNS, start=1):
                ws.cell(row=1, column=idx, value=name)
        else:
            # Add any missing columns at the end if the workbook differs
            existing = [safe_strip(h) for h in headers]
            for col_name in EXPECTED_COLUMNS:
                if col_name not in existing:
                    ws.cell(row=1, column=ws.max_column + 1, value=col_name)

    # Rebuild header map after possible changes
    headers = [cell.value for cell in ws[1]]
    header_map = {str(name).strip(): idx for idx, name in enumerate(headers, start=1) if name}

    return wb, ws, header_map


def append_row_to_excel(
    ws,
    header_map: Dict[str, int],
    company_name: Optional[str],
    email: Optional[str],
    address: Optional[str],
    phone: Optional[str],
    website: Optional[str],
):
    new_row_idx = ws.max_row + 1

    row_values = {
        "Company name": company_name,
        "Email": email,
        "First name": None,
        "Last name": None,
        "Job title": None,
        "country/address": address,
        "Phone": phone,
        "Website": website,
        "Industry type": None,
    }

    for col_name, value in row_values.items():
        col_idx = header_map.get(col_name)
        if col_idx:
            ws.cell(row=new_row_idx, column=col_idx, value=value)


def main():
    items = load_input_json(JSON_FILE)
    wb, ws, header_map = open_or_create_workbook(XLSX_FILE)

    total = len(items)
    print(f"[INFO] Loaded {total} items from {JSON_FILE}")

    for i, item in enumerate(items, start=1):
        winery_id = safe_strip(item.get("id"))
        if not winery_id:
            print(f"[WARN] Skipping item #{i}: missing id")
            continue

        print(f"[INFO] ({i}/{total}) Processing {winery_id}")

        details = fetch_winery_data(winery_id)

        append_row_to_excel(
            ws=ws,
            header_map=header_map,
            company_name=details.get("company_name"),
            email=details.get("email"),
            address=details.get("address"),
            phone=details.get("phone"),
            website=details.get("website"),
        )

        # Save after each append so progress is not lost
        wb.save(XLSX_FILE)
        time.sleep(SLEEP_BETWEEN_REQUESTS)

    print(f"[DONE] Finished. Data appended to {XLSX_FILE}")


if __name__ == "__main__":
    main()