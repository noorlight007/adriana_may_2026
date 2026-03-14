import os
from dotenv import load_dotenv
load_dotenv()
import re
import json
from typing import Any, Dict

from openai import OpenAI
from jsonschema import Draft202012Validator
from openpyxl import load_workbook


MODEL = "gpt-4o-search-preview-2025-03-11"
EXCEL_FILE = "francewines.xlsx"


def return_Schema() -> Dict[str, Any]:
    SCHEMA: Dict[str, Any] = {
        "type": "object",
        "additionalProperties": False,
        "properties": {
            "email": {
                "type": "string",
                "description": 'Company or contact email. Return "null" if not found.'
            }
        },
        "required": ["email"]
    }
    return SCHEMA


def get_all_instructions():
    SYSTEM_INSTRUCTIONS = (
        "You are an information extraction engine.\n"
        "Your task is to identify and extract only one email address from the provided website URL.\n"
        "- Extract the most relevant company email or contact person's work email.\n"
        "- Only extract an email if it is explicitly available from the website.\n"
        "- Do NOT invent, guess, or infer any email address.\n"
        "- If no valid email is found, return the literal string \"null\".\n"
        "- Return ONLY raw JSON matching the required schema.\n"
        "- Do not return markdown fences, explanations, or extra text.\n"
        "- Output format must be exactly like:\n"
        "{\n"
        "  \"email\": \"example@company.com\"\n"
        "}\n"
        "- If no email exists, output:\n"
        "{\n"
        "  \"email\": \"null\"\n"
        "}\n"
    )

    USER_INSTRUCTIONS_TEMPLATE = (
        "Extract the single most relevant email address from the following website.\n"
        "If no email is present, return \"null\".\n"
        "Return ONLY raw JSON matching the required schema.\n\n"
        "WEBSITE URL START\n"
        "{website_url}\n"
        "WEBSITE URL END\n"
    )

    return SYSTEM_INSTRUCTIONS, USER_INSTRUCTIONS_TEMPLATE


def strip_code_fences(s: str) -> str:
    """Remove markdown code fences like ``` or ```json."""
    s = s.strip()
    if s.startswith("```"):
        lines = s.splitlines()
        if lines and lines[0].startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        s = "\n".join(lines).strip()
    return s


def find_json_block(s: str) -> str:
    """If JSON isn't clean, try to extract the largest {...} or [...] block."""
    s = s.strip()
    obj_match = re.search(r"\{.*\}\s*$", s, flags=re.DOTALL)
    if obj_match:
        return obj_match.group(0)
    arr_match = re.search(r"\[.*\]\s*$", s, flags=re.DOTALL)
    if arr_match:
        return arr_match.group(0)
    return s


def normalize_nulls(data: Dict[str, Any]) -> Dict[str, Any]:
    """Force missing/empty email to the literal string 'null'."""
    def to_str_null(v: Any) -> str:
        if v is None:
            return "null"
        if isinstance(v, str):
            return v.strip() if v.strip() else "null"
        return str(v)

    data["email"] = to_str_null(data.get("email"))
    return data


def get_header_indexes(ws) -> Dict[str, int]:
    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is not None:
            headers[str(value).strip()] = col
    return headers


def extract_email_from_website(client: OpenAI, website_url: str) -> str:
    SCHEMA = return_Schema()
    SYSTEM_INSTRUCTIONS, USER_INSTRUCTIONS_TEMPLATE = get_all_instructions()

    messages = [
        {"role": "system", "content": SYSTEM_INSTRUCTIONS},
        {
            "role": "user",
            "content": USER_INSTRUCTIONS_TEMPLATE.format(website_url=website_url),
        },
    ]

    resp = client.chat.completions.create(
        model=MODEL,
        messages=messages,
    )

    raw = resp.choices[0].message.content or ""
    content = strip_code_fences(raw)

    try:
        data = json.loads(content)
    except json.JSONDecodeError:
        candidate = find_json_block(content)
        data = json.loads(candidate)

    Draft202012Validator(SCHEMA).validate(data)
    data = normalize_nulls(data)

    return data["email"]


def main():
    api_key = os.getenv("OPENAI_API")
    if not api_key:
        raise ValueError("OPENAI_API environment variable is not set.")

    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"{EXCEL_FILE} not found.")

    client = OpenAI(api_key=api_key)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    headers = get_header_indexes(ws)

    if "Website" not in headers:
        raise ValueError('Column "Website" not found in francewines.xlsx')
    if "Email" not in headers:
        raise ValueError('Column "Email" not found in francewines.xlsx')

    website_col = headers["Website"]
    email_col = headers["Email"]
    i = 0
    for row in range(2, ws.max_row + 1):
        i+= 1
        website_value = ws.cell(row=row, column=website_col).value
        current_email = ws.cell(row=row, column=email_col).value

        website_url = str(website_value).strip() if website_value else ""
        existing_email = str(current_email).strip() if current_email else ""

        if not website_url:
            print(f"Row {row}: No website found. Skipping.")
            continue

        if existing_email:
            print(f"Row {row}: Email already exists ({existing_email}). Skipping.")
            continue

        try:
            email = extract_email_from_website(client, website_url)

            if email != "null":
                ws.cell(row=row, column=email_col).value = email
                wb.save(EXCEL_FILE)
                print(f"Row {row}: Email written -> {email}")
            else:
                print(f"Row {row}: No email found for {website_url}")

        except Exception as e:
            print(f"Row {row}: Error processing {website_url}: {e}")
            print("--------------------------------------------------")
            continue

        if i == 5:
            print("Processed 5 rows, stopping to avoid hitting rate limits.")
            break

    wb.save(EXCEL_FILE)
    wb.close()
    print("Done.")


if __name__ == "__main__":
    main()
