import sys
from openpyxl import load_workbook


def remove_next_duplicate_emails(file_path: str, sheet_name: str = None, output_path: str = None) -> None:
    # Load workbook
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    # Read header row
    headers = [cell.value for cell in sheet[1]]

    if "Email" not in headers:
        print('Error: "Email" column not found.')
        return

    email_col_index = headers.index("Email") + 1  # openpyxl is 1-based

    seen_emails = set()
    cleared_rows = []

    # Keep first occurrence, clear later duplicates
    for row_num in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row_num, column=email_col_index)
        cell_value = cell.value

        if cell_value is None:
            continue

        email = str(cell_value).strip().lower()
        if not email:
            continue

        if email in seen_emails:
            cell.value = ""
            cleared_rows.append(row_num)
        else:
            seen_emails.add(email)

    # Save workbook
    save_path = output_path if output_path else file_path
    workbook.save(save_path)

    if cleared_rows:
        print(f"Done. Cleared duplicate emails in rows: {cleared_rows}")
    else:
        print("No duplicate emails found.")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python remove_next_duplicate_emails.py <xlsx_file_path> [sheet_name] [output_path]")
        sys.exit(1)

    file_path = sys.argv[1]
    sheet_name = sys.argv[2] if len(sys.argv) > 2 else None
    output_path = sys.argv[3] if len(sys.argv) > 3 else None

    remove_next_duplicate_emails(file_path, sheet_name, output_path)