import json
from openpyxl import load_workbook

JSON_FILE = "decanter_italy.json"
XLSX_FILE = "awards_decanter_Italy.xlsx"
TARGET_HEADER = "Company Name"


def main():
    # Load JSON data
    with open(JSON_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Open workbook and select active sheet
    wb = load_workbook(XLSX_FILE)
    ws = wb.active

    # Find the "Company Name" column by header row
    company_col = None
    for cell in ws[1]:
        if cell.value == TARGET_HEADER:
            company_col = cell.column
            break

    if company_col is None:
        raise ValueError(f'Column header "{TARGET_HEADER}" not found in the first row.')

    # Write each JSON item's "name" into the Company Name column starting from row 2
    for row_index, item in enumerate(data, start=2):
        ws.cell(row=row_index, column=company_col, value=item.get("name", ""))

    # Save changes
    wb.save(XLSX_FILE)
    print(f'Successfully wrote {len(data)} company names into "{XLSX_FILE}".')


if __name__ == "__main__":
    main()