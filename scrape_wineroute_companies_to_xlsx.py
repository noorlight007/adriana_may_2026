#!/usr/bin/env python3
import time
from typing import List, Set, Tuple
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

BASE_LISTING_URL = "https://www.wineroute.alsace/wineries/"
OUTPUT_XLSX = "wineroute_wineries.xlsx"
REQUEST_DELAY_SECONDS = 0.5


def fetch_html(url: str, timeout: int = 30) -> str:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (X11; Linux x86_64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/123.0 Safari/537.36"
        )
    }
    response = requests.get(url, headers=headers, timeout=timeout)
    response.raise_for_status()
    return response.text


def get_listing_page_url(page_number: int) -> str:
    if page_number == 1:
        return BASE_LISTING_URL
    return f"{BASE_LISTING_URL}page/{page_number}/"


def extract_winery_links(listing_html: str, base_url: str) -> List[str]:
    soup = BeautifulSoup(listing_html, "html.parser")

    # Preferred selector based on the provided HTML snippet.
    container = soup.select_one("div.wrapper-cards.wrapper.js-listing-card-container")
    if container is None:
        container = soup

    links: List[str] = []
    seen: Set[str] = set()

    for a_tag in container.select("a.cards-v2.card-sit[href]"):
        href = (a_tag.get("href") or "").strip()
        if not href:
            continue
        absolute = urljoin(base_url, href)
        if absolute not in seen:
            seen.add(absolute)
            links.append(absolute)

    return links


def collect_all_winery_links() -> List[str]:
    all_links: List[str] = []
    seen_links: Set[str] = set()
    empty_streak = 0
    page_number = 1

    while True:
        url = get_listing_page_url(page_number)
        print(f"Scanning listing page {page_number}: {url}")

        try:
            html = fetch_html(url)
        except requests.HTTPError as exc:
            status_code = getattr(exc.response, "status_code", None)
            if status_code == 404:
                print(f"Stopping at page {page_number}: page not found.")
                break
            raise

        page_links = extract_winery_links(html, url)

        if not page_links:
            empty_streak += 1
            print(f"No winery links found on page {page_number}.")
            if empty_streak >= 2:
                break
        else:
            empty_streak = 0

        new_links = 0
        for link in page_links:
            if link not in seen_links:
                seen_links.add(link)
                all_links.append(link)
                new_links += 1

        print(f"Found {len(page_links)} links on page {page_number}, {new_links} new.")

        # Stop if a page contains no new links.
        if page_links and new_links == 0:
            break

        page_number += 1
        time.sleep(REQUEST_DELAY_SECONDS)

    return all_links


def clean_text(value: str) -> str:
    return " ".join(value.split()).strip()


def scrape_winery_detail(url: str) -> Tuple[str, str, str, str]:
    html = fetch_html(url)
    soup = BeautifulSoup(html, "html.parser")

    company = ""
    address = ""
    phone = ""
    website = ""

    # Company: H1 inside the detail page header area.
    h1 = soup.select_one("div.content-wrapper.caption.aligncenter h1") or soup.find("h1")
    if h1:
        company = clean_text(h1.get_text(" ", strip=True))

    # Address: first p + second p inside div.address.
    address_box = soup.select_one("div.address")
    if address_box:
        p_tags = address_box.find_all("p")
        if len(p_tags) >= 2:
            address = clean_text(
                f"{p_tags[0].get_text(' ', strip=True)} {p_tags[1].get_text(' ', strip=True)}"
            )
        elif len(p_tags) == 1:
            address = clean_text(p_tags[0].get_text(" ", strip=True))

    # Phone: href value from tel link, removing tel:
    phone_link = soup.select_one('div.cta a.gaevent-tel[href^="tel:"]')
    if phone_link and phone_link.get("href"):
        phone = phone_link["href"].replace("tel:", "", 1).strip()

    # Website: href from website icon link.
    website_link = soup.select_one('a.gaevent-site[href]')
    if website_link and website_link.get("href"):
        website = website_link["href"].strip()

    return company, address, phone, website


def init_or_load_workbook(path: str):
    try:
        workbook = load_workbook(path)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Wineries"
        sheet.append(["Company", "Address", "Phone", "Website"])
        workbook.save(path)
    return workbook, sheet


def main() -> None:
    workbook, sheet = init_or_load_workbook(OUTPUT_XLSX)

    winery_links = collect_all_winery_links()
    print(f"Total winery detail pages collected: {len(winery_links)}")

    for index, winery_url in enumerate(winery_links, start=1):
        print(f"[{index}/{len(winery_links)}] Scraping: {winery_url}")
        try:
            company, address, phone, website = scrape_winery_detail(winery_url)
        except Exception as exc:
            print(f"Failed to scrape {winery_url}: {exc}")
            company, address, phone, website = "", "", "", ""

        # Append one row, save immediately, then continue.
        sheet.append([company, address, phone, website])
        workbook.save(OUTPUT_XLSX)

        time.sleep(REQUEST_DELAY_SECONDS)

    print(f"Done. Spreadsheet saved to {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
