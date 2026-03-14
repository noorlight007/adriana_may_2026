import json
import os
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook


BASE_URL = "https://www.bourgogne-wines.com"  # Change this to the real website domain
HREF_JSON_FILE = "href_list.json"
EXCEL_FILE = "france_wines.xlsx"


def clean_text(text):
    if not text:
        return ""
    return " ".join(text.split()).strip()


def split_name(full_name):
    full_name = clean_text(full_name)
    if not full_name:
        return "", ""
    parts = full_name.split()
    first_name = parts[0]
    last_name = " ".join(parts[1:]) if len(parts) > 1 else ""
    return first_name, last_name


def get_li_value_from_icon(info_box, icon_name):
    """
    Find the <li> whose <img src> contains the given icon file name,
    then return its text content.
    """
    for li in info_box.find_all("li"):
        img = li.find("img")
        if img:
            src = img.get("src", "")
            if icon_name in src:
                return clean_text(li.get_text(" ", strip=True))
    return ""


def get_website(info_box):
    for li in info_box.find_all("li"):
        img = li.find("img")
        if img:
            src = img.get("src", "")
            if "icon-link.svg" in src:
                a_tag = li.find("a", href=True)
                if a_tag:
                    return clean_text(a_tag["href"])
                return clean_text(li.get_text(" ", strip=True))
    return ""


def scrape_page(url):
    print(f"Scraping: {url}")
    try:
        response = requests.get(url, timeout=30)
        response.raise_for_status()
    except Exception as e:
        print(f"Failed to fetch {url}: {e}")
        return {
            "Company name": "",
            "Email": "",
            "First name": "",
            "Last name": "",
            "Job title": "",
            "country/address": "",
            "Phone": "",
            "Website": "",
        }

    soup = BeautifulSoup(response.text, "html.parser")
    info_box = soup.find("div", class_="domain-infos")

    if not info_box:
        print(f"No domain-infos section found in: {url}")
        return {
            "Company name": "",
            "Email": "",
            "First name": "",
            "Last name": "",
            "Job title": "",
            "country/address": "",
            "Phone": "",
            "Website": "",
        }

    company_name = ""
    company_tag = info_box.find("strong", class_="domain-name")
    if company_tag:
        company_name = clean_text(company_tag.get_text())

    address = get_li_value_from_icon(info_box, "icon-address.svg")
    person_name = get_li_value_from_icon(info_box, "icon-name.svg")
    phone = get_li_value_from_icon(info_box, "icon-phone.svg")  # first matching phone only
    website = get_website(info_box)

    first_name, last_name = split_name(person_name)

    return {
        "Company name": company_name,
        "Email": "",
        "First name": first_name,
        "Last name": last_name,
        "Job title": "",
        "country/address": address,
        "Phone": phone,
        "Website": website,
    }


def load_links():
    with open(HREF_JSON_FILE, "r", encoding="utf-8") as f:
        hrefs = json.load(f)

    full_urls = [urljoin(BASE_URL, href) for href in hrefs]
    return full_urls


def find_header_indexes(ws):
    headers = {}
    for col in range(1, ws.max_column + 1):
        header_value = ws.cell(row=1, column=col).value
        if header_value:
            headers[str(header_value).strip()] = col
    return headers


def append_data_to_excel(data_rows):
    if not os.path.exists(EXCEL_FILE):
        raise FileNotFoundError(f"{EXCEL_FILE} was not found.")

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    headers = find_header_indexes(ws)

    required_headers = [
        "Company name",
        "Email",
        "First name",
        "Last name",
        "Job title",
        "country/address",
        "Phone",
        "Website",
    ]

    for header in required_headers:
        if header not in headers:
            raise ValueError(f'Missing required column header: "{header}"')

    next_row = ws.max_row + 1

    for row_data in data_rows:
        ws.cell(row=next_row, column=headers["Company name"], value=row_data["Company name"])
        ws.cell(row=next_row, column=headers["Email"], value=row_data["Email"])
        ws.cell(row=next_row, column=headers["First name"], value=row_data["First name"])
        ws.cell(row=next_row, column=headers["Last name"], value=row_data["Last name"])
        ws.cell(row=next_row, column=headers["Job title"], value=row_data["Job title"])
        ws.cell(row=next_row, column=headers["country/address"], value=row_data["country/address"])
        ws.cell(row=next_row, column=headers["Phone"], value=row_data["Phone"])
        ws.cell(row=next_row, column=headers["Website"], value=row_data["Website"])
        next_row += 1

    wb.save(EXCEL_FILE)
    print(f"Data written successfully to {EXCEL_FILE}")


def main():
    links = load_links()
    all_data = []

    for link in links:
        data = scrape_page(link)
        all_data.append(data)

    append_data_to_excel(all_data)
    print("Done.")


if __name__ == "__main__":
    main()